import pathlib
import shutil
import openpyxl as xl
import os
import vars
import fee
import pandas as pd
import traceback
import sys, os


#print(traceback.format_exc())

def transformFile(file: str,qtr: str,year: str)->list[str]:
    '''
    Fucntion for individual excel files in the folder. 
    Get new file path & file name
    Replace new path & name with existing path & name
    '''
    file_info = getnewfilename(file,qtr,year)

    if file_info == []:
        print("An error with the file has been detected. Ignoring file.")
        return []

    prefix = file_info[0] #venture name, partner venture name if not in vars list
    suffix = file_info[1] #type of report
    status = file_info[2] #Not Found, Mult-Venture,Single-Venture

    newfilepath = getnewfilepath(file,qtr,year)

    if status == "Not Found":
        return [newfilepath,prefix,suffix]
    
    if status == "Single-Venture":
        rename_file(file,newfilepath,prefix,suffix)
        return []

    elif status == "Multi-Venture":
        rename_file(file,newfilepath,prefix,suffix+" - Mutli-Venture")
        return []

    else:
        print("Status Error. Check getnewfilename function.")
        return []
 
def getnewfilepath(file: str,qtr: str,year: str)->str:
    '''
    get file path of new destination
    '''
    # target_dir = (
    #     str(pathlib.Path.home()) + "/OneDrive - Quadreal Property Group" + od_path
    # ) #real path
    # target_dir = (
    #     str(pathlib.Path.home()) + "/OneDrive - Quadreal Property Group" + vars.test_path
    # ) #work test path

    target_dir = vars.home_test_path #home test path

    newFolderName = year + " Q" + qtr
    # print("quarter is " + qtr)
    path = os.path.join(target_dir,newFolderName).replace("\\","/")
    
    if not os.path.exists(path):    
        os.mkdir(path)
    else:
        print("Target directory already exists.")
    return path


def getnewfilename(file: str,qtr: str,year: str)->list[str]:
    '''
    Get venture name and type of report (AM Fee or Deployment forecast or both) 
        - If neither report exists, return empty
        - If only AM Fee and venture name exists, record fee and create new file name
        - If only Deployment Forecast and venture name exists, create new file name
        - If either/both exist but no name or wrong name, flag error and provide new name in UI
    '''
    try:
        wb = xl.load_workbook(filename = file,data_only=True)   
    except:
        print("Probably not an excel file. Ignoring this file...")
        return []

    
    deployment = False # this means deployment sheet exists, but not necesarily the correct venture name
    am_fee = False # this means AM fee sheet exists, but not necessarily with the correct venture name
    QR_venture_name = ""
    venture_count = 0

    '''
    Checks each worksheet to see if capital deployment sheet or AM fee sheet exist
    '''
    for ws in wb:
        am_ws = ""
        deployment_sheet = False
        if ws.sheet_state == "visible":
            venture_name = ""
            if str(ws.cell(row=4,column=1).value) == "INVESTMENT NAME:":
                # print("capital deployment sheet is visible!!")
                if str(ws.cell(row=4,column=2).value) != "":
                    venture_name = str(ws.cell(row=4,column=2).value).strip()
                deployment = True
                deployment_sheet = True
            elif str(ws.cell(row=6,column=1).value) == "INVESTMENT NAME:":
                if str(ws.cell(row=6,column=2).value) != "":
                    venture_name = str(ws.cell(row=6,column=2).value).strip()
                am_fee = True
                am_ws = ws.title

        qr_name_temp = ""
        if not deployment_sheet and not am_ws == "":
            print("   Not one of the expected reports. Trying next worksheet..")
            continue
        else:
            qr_name_temp = vars.venture_names.get(venture_name.lower())
            if qr_name_temp is not None:
                QR_venture_name = str(qr_name_temp).strip()
                print(f'{venture_name} has been converted to {QR_venture_name}')
                venture_count += 1
                if am_ws != "":
                    try:
                        fee.recordfee(QR_venture_name,file,am_ws,qtr,year)
                    except Exception as e:
                        print(traceback.format_exc())
                        print(e)
            else:
                QR_venture_name = venture_name
                print("Venture name not found. Wrong name or cell is empty")

    print("No more worksheets to look through")
    
    '''
    If no relevant worksheet is found, leave the file alone & flag
    If worksheet is found, get the name of the venture using the cell found above
    '''
    suffix = " - Q" + str(qtr)
    if deployment & am_fee:
        suffix += " QRI Capital Deployment Forecast and AM Fee.xlsx"
    elif deployment:
        suffix += " QRI Capital Deployment Forecast.xlsx"
    elif am_fee:
        suffix += " QRI AM Fee.xlsx"
    else:
        suffix += ""

    prefix = QR_venture_name

    if venture_count == 0:
        vn_status = "Not Found"
        print("No venture name has been mapped in this file")
    elif venture_count > 1:
        vn_status = "Multi-Venture"
        print("There are reports from multiple ventures in this file")
    else:
        vn_status = "Single-Venture"
        print("This file is solely for " + QR_venture_name)

    return [prefix,suffix,vn_status]


def exception_ventures():
    print("Venture Exception handling not implemented yet, exiting now...")


def rename_file(file: str,newfilepath: str,prefix: str,suffix:str):

    newfile = os.path.join(newfilepath,prefix+suffix).replace("\\","/")
    shutil.move(file, newfile)